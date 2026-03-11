import streamlit as st
import openpyxl
from datetime import datetime, date, timedelta
import io
import smtplib
from email.mime.multipart import MIMEMultipart
from email.mime.base import MIMEBase
from email.mime.text import MIMEText
from email import encoders
import os
import base64

# ─── Page Config ──────────────────────────────────────────────────────────────
st.set_page_config(
    page_title="FANUC FIR Generator",
    page_icon="🤖",
    layout="centered",
    initial_sidebar_state="collapsed"
)

# ══════════════════════════════════════════════════════════════════════════════
# AD CONFIGURATION
# ══════════════════════════════════════════════════════════════════════════════
SHOW_ADS = False
ADSENSE_PUBLISHER_ID = "ca-pub-XXXXXXXXXXXXXXXX"
ADSENSE_SLOT_TOP    = "1111111111"
ADSENSE_SLOT_MID    = "2222222222"
ADSENSE_SLOT_BOTTOM = "3333333333"
# ══════════════════════════════════════════════════════════════════════════════

def inject_adsense():
    if not SHOW_ADS: return
    st.markdown(
        f'<script async src="https://pagead2.googlesyndication.com/pagead/js/adsbygoogle.js'
        f'?client={ADSENSE_PUBLISHER_ID}" crossorigin="anonymous"></script>',
        unsafe_allow_html=True)

def ad_slot(slot_id, height="90px"):
    if not SHOW_ADS: return
    st.markdown(f"""
    <div style="text-align:center;margin:10px 0;">
      <ins class="adsbygoogle" style="display:block;width:100%;height:{height}"
           data-ad-client="{ADSENSE_PUBLISHER_ID}" data-ad-slot="{slot_id}"
           data-ad-format="auto" data-full-width-responsive="true"></ins>
      <script>(adsbygoogle = window.adsbygoogle || []).push({{}});</script>
    </div>""", unsafe_allow_html=True)

inject_adsense()

# ─── FANUC Logo as inline SVG (yellow bg, red bold text — matches real logo) ──
FANUC_LOGO_SVG = """
<svg xmlns="http://www.w3.org/2000/svg" viewBox="0 0 200 60" width="110" height="33">
  <rect width="200" height="60" fill="#FFD700"/>
  <text x="100" y="46" font-family="Arial Black, Arial" font-weight="900"
        font-size="42" fill="#CC0000" text-anchor="middle" letter-spacing="-1">FANUC</text>
</svg>
"""

# ─── Light-Theme Mobile CSS ───────────────────────────────────────────────────
st.markdown("""
<style>
@import url('https://fonts.googleapis.com/css2?family=Inter:wght@400;500;600;700&display=swap');

:root {
  --bg:       #F0F2F5;
  --white:    #FFFFFF;
  --border:   #E0E3E8;
  --txt:      #1A1A2E;
  --txt2:     #6B7280;
  --yellow:   #FFD700;
  --red:      #CC0000;
  --accent:   #1A1A2E;
  --radius:   14px;
  --shadow:   0 2px 12px rgba(0,0,0,0.07);
}

html, body, [class*="css"] {
  font-family: 'Inter', sans-serif !important;
  background: var(--bg) !important;
  color: var(--txt) !important;
}

#MainMenu, footer, header { visibility: hidden; }

.block-container {
  padding: 0 0 80px !important;
  max-width: 480px !important;
  margin: 0 auto !important;
}

/* ── Top bar ── */
.topbar {
  background: #1A1A2E;
  padding: 0 16px;
  margin: 0 0 0;
  display: flex;
  align-items: center;
  justify-content: space-between;
  height: 58px;
  position: sticky;
  top: 0;
  z-index: 999;
  box-shadow: 0 2px 8px rgba(0,0,0,0.25);
}
.topbar-right {
  background: var(--yellow);
  color: #000;
  font-weight: 700;
  font-size: 13px;
  letter-spacing: 1px;
  border: none;
  border-radius: 8px;
  padding: 9px 16px;
  cursor: pointer;
  display: flex;
  align-items: center;
  gap: 6px;
  white-space: nowrap;
}

/* ── Step pills ── */
.steps-wrap {
  background: var(--white);
  border-bottom: 1px solid var(--border);
  padding: 10px 14px;
  margin-bottom: 0;
}
.steps {
  display: flex; gap: 6px; overflow-x: auto;
  scrollbar-width: none;
}
.steps::-webkit-scrollbar { display: none; }
.sp {
  flex-shrink: 0; border-radius: 20px; padding: 6px 14px;
  font-size: 12px; font-weight: 600; letter-spacing: 0.5px;
  white-space: nowrap; border: 1.5px solid var(--border);
  background: var(--white); color: var(--txt2); cursor: default;
}
.sp.on  { background: #1A1A2E; color: #fff; border-color: #1A1A2E; }
.sp.done { background: #F0FFF4; color: #16A34A; border-color: #86EFAC; }

/* ── Page title inside content ── */
.page-title {
  padding: 22px 16px 6px;
}
.page-title .icon {
  font-size: 32px; display: block; margin-bottom: 6px;
}
.page-title h2 {
  font-size: 26px; font-weight: 700; color: var(--txt);
  margin: 0 0 2px; line-height: 1.2;
}
.page-title p {
  font-size: 13px; color: var(--txt2); margin: 0;
}

/* ── White card ── */
.card {
  background: var(--white);
  border-radius: var(--radius);
  box-shadow: var(--shadow);
  padding: 20px 16px;
  margin: 12px 12px;
}
.card-label {
  font-size: 11px; font-weight: 700; color: var(--txt2);
  letter-spacing: 1.5px; text-transform: uppercase;
  margin-bottom: 14px;
  display: flex; align-items: center; gap: 8px;
}
.card-label::after {
  content: ''; flex: 1; height: 1px; background: var(--border);
}

/* ── Filename badge ── */
.fname {
  background: #FFFBEB;
  border: 1.5px solid var(--yellow);
  border-radius: 10px;
  padding: 12px 14px;
  font-size: 13px; font-weight: 700;
  color: #92400E;
  text-align: center;
  word-break: break-all;
  margin: 12px 0 4px;
  letter-spacing: 0.3px;
}
.fname-label {
  font-size: 10px; color: var(--txt2); text-align: center;
  letter-spacing: 1px; text-transform: uppercase; margin-bottom: 4px;
}

/* ── Summary row ── */
.summary-row {
  display: grid; grid-template-columns: repeat(4,1fr);
  gap: 8px; margin: 12px 12px;
}
.summary-cell {
  background: var(--white); border-radius: 10px;
  box-shadow: var(--shadow); padding: 10px 6px;
  text-align: center;
}
.summary-cell .val {
  font-size: 18px; font-weight: 700; color: #1A1A2E;
  font-variant-numeric: tabular-nums;
}
.summary-cell .lbl {
  font-size: 10px; color: var(--txt2);
  letter-spacing: 1px; text-transform: uppercase; margin-top: 2px;
}

/* ── Inputs — light theme, 16px to prevent iOS zoom ── */
.stTextInput > div > div > input,
.stSelectbox > div > div,
.stDateInput > div > div > input,
.stNumberInput > div > div > input,
.stTextArea textarea {
  background: #F9FAFB !important;
  border: 1.5px solid var(--border) !important;
  border-radius: 10px !important;
  color: var(--txt) !important;
  font-size: 16px !important;
  min-height: 50px !important;
  box-shadow: none !important;
}
.stTextInput > div > div > input:focus,
.stTextArea textarea:focus {
  border-color: #1A1A2E !important;
  background: #fff !important;
  box-shadow: 0 0 0 2px rgba(26,26,46,0.08) !important;
}

/* Labels */
.stTextInput label, .stSelectbox label, .stDateInput label,
.stNumberInput label, .stTextArea label, .stCheckbox label {
  color: var(--txt2) !important;
  font-size: 11px !important;
  font-weight: 700 !important;
  letter-spacing: 1.2px !important;
  text-transform: uppercase !important;
}

/* ── Buttons ── */
.stButton > button {
  width: 100% !important;
  min-height: 52px !important;
  background: #1A1A2E !important;
  color: #fff !important;
  font-family: 'Inter', sans-serif !important;
  font-weight: 700 !important;
  font-size: 15px !important;
  letter-spacing: 1.5px !important;
  border: none !important;
  border-radius: 12px !important;
  margin: 4px 0 !important;
  transition: opacity 0.15s !important;
}
.stButton > button:hover { opacity: 0.88 !important; }

/* Primary yellow action button — first button on generate page */
.stDownloadButton > button {
  width: 100% !important;
  min-height: 52px !important;
  background: var(--yellow) !important;
  color: #000 !important;
  font-family: 'Inter', sans-serif !important;
  font-weight: 700 !important;
  font-size: 15px !important;
  letter-spacing: 1.5px !important;
  border: none !important;
  border-radius: 12px !important;
}

/* Back button — secondary style */
button[kind="secondary"] {
  background: #F3F4F6 !important;
  color: #374151 !important;
}

/* ── Expander (day blocks) ── */
.streamlit-expanderHeader {
  background: var(--white) !important;
  border: 1.5px solid var(--border) !important;
  border-radius: 10px !important;
  min-height: 52px !important;
  color: var(--txt) !important;
  font-size: 15px !important;
  font-weight: 600 !important;
  margin-bottom: 6px !important;
}
.streamlit-expanderContent {
  background: #FAFAFA !important;
  border: 1.5px solid var(--border) !important;
  border-top: none !important;
  border-radius: 0 0 10px 10px !important;
  padding: 14px !important;
}

/* ── Checkbox ── */
.stCheckbox > label { min-height: 40px !important; font-size: 14px !important; }

/* ── Metrics ── */
[data-testid="metric-container"] {
  background: var(--white) !important;
  border: 1.5px solid var(--border) !important;
  border-radius: 10px !important;
  padding: 10px 8px !important;
}
[data-testid="stMetricValue"] { color: #1A1A2E !important; font-size: 20px !important; font-weight: 700 !important; }
[data-testid="stMetricLabel"] { color: var(--txt2) !important; font-size: 10px !important; letter-spacing: 1px !important; }

hr { border-color: var(--border) !important; margin: 14px 0 !important; }
.stAlert { border-radius: 10px !important; font-size: 14px !important; }
::-webkit-scrollbar { width: 4px; }
::-webkit-scrollbar-thumb { background: var(--border); border-radius: 2px; }

/* ── Bottom nav spacing ── */
.bottom-nav { padding: 0 12px; margin-top: 8px; }
</style>
""", unsafe_allow_html=True)


# ─── Constants & helpers ──────────────────────────────────────────────────────
REPORT_TYPES = {"Installation (I)": "I", "Production Support (P)": "P", "Other (O)": "O"}

def generate_filename(project_num, start_date, emp_no, last_name, report_type):
    """Use emp_no if available, otherwise fall back to last_name."""
    if not all([project_num, start_date]):
        return "FIR-______-______-___(_).xlsm"
    identifier = emp_no.strip() if emp_no and emp_no.strip() else (last_name.strip() if last_name and last_name.strip() else "___")
    suffix = REPORT_TYPES.get(report_type, "O")
    return f"FIR-{project_num}-{start_date.strftime('%y%m%d')}-{identifier}({suffix}).xlsm"


def populate_fir_excel(data):
    template_path = os.path.join(os.path.dirname(__file__), "FIR_template.xlsm")
    wb = openpyxl.load_workbook(template_path, keep_vba=True)
    ws  = wb["Field Installation"]
    ws2 = wb["SUMFIR"]

    # ── Row 2: Project # and Program/Process/Cell ──────────────────────────────
    # Project # → I2  (merged I2:K2;  label = G2:H2 "Project #:")
    ws["I2"] = int(data["project_num"]) if data["project_num"].isdigit() else data["project_num"]
    # Program / Process / Cell → O2  (merged O2:AA2;  label = L2:N2)
    ws["O2"] = data.get("program_process_cell", "")

    # ── Row 3: Customer Contact, Phone, Plant/Location ─────────────────────────
    # Customer Contact Name → D3  (merged D3:H3;  label = A3:C3)
    ws["D3"] = data.get("customer_contact", "")
    # Phone # → L3  (merged L3:O3;  label = J3:K3)
    ws["L3"] = data.get("phone", "")
    # Plant / Location → S3  (merged S3:AA3;  label = P3:R3)
    ws["S3"] = data.get("plant_location", "")

    # ── Row 4: Type of Report (radio → SUMFIR!AV2: 1=I, 2=P, 3=O) ────────────
    # The radio buttons are form controls linked to SUMFIR!AV2
    # We write directly to SUMFIR AV2 (col 48) to set the report type
    report_val = {"Installation (I)": 1, "Production Support (P)": 2, "Other (O)": 3}
    ws2.cell(row=2, column=48).value = report_val.get(data["report_type"], 3)

    # Exp Rep Yes/No → SUMFIR!AW2 (col 49): 1=Yes, 2=No
    ws2.cell(row=2, column=49).value = 1 if data.get("exp_rep", False) else 2

    # ── Row 5: Expense Amount ──────────────────────────────────────────────────
    # Expense Amount → X5  (merged X5:AA5;  label = V5:W5)
    ws["X5"] = data.get("expense_amount", 0) or 0

    # ── Row 55/56: Engineer signature block ────────────────────────────────────
    # Engineer first name → E55  (merged E55:J55)
    ws["E55"] = data.get("engineer_first_name", "")
    # Emp # OR Last Name → M55  (merged M55:N55;  label = K55:L55 "EmpNo: / Last Name:")
    emp = (data.get("emp_no") or "").strip()
    ws["M55"] = emp if emp else (data.get("last_name") or "").strip()
    # Print name → B56  (merged B56:I56;  label = A56 "Print:")
    ws["B56"] = data.get("engineer_print_name", "")

    # ── Day entries ────────────────────────────────────────────────────────────
    # Confirmed layout per day block (7 days):
    #   date_row  : A col — day1=6 input, days2-7 carry formula =prev+1
    #   work_row  : F col (col6) — work description text (merged F6:Z10 etc.)
    #   hours_row : S=col1, OT=col2, DT=col3, TT=col4, W=col5, Wait=AA(col27)
    #               day1=row9, day2=15, day3=21, day4=27, day5=33, day6=39, day7=45
    day_map = [
        # (date_row, hours_row, work_row)
        (6,   9,  6),
        (12, 15, 12),
        (18, 21, 18),
        (24, 27, 24),
        (30, 33, 30),
        (36, 39, 36),
        (42, 45, 42),
    ]
    for i, (dr, hr, wr) in enumerate(day_map):
        d = data["days"][i] if i < len(data["days"]) else {}
        if not d.get("active"):
            continue
        # Only day 1 date is a plain input cell; days 2-7 auto-increment via formula
        if i == 0 and data.get("start_date"):
            ws.cell(row=dr, column=1).value = datetime.combine(
                data["start_date"], datetime.min.time()
            )
        # Hours columns: S=1, OT=2, DT=3, TT=4, W=5, Wait=27(AA)
        for col, key in [(1,"straight"),(2,"overtime"),(3,"doubletime"),
                         (4,"travel_time"),(5,"working"),(27,"wait")]:
            val = d.get(key)
            ws.cell(row=hr, column=col).value = val if val else None
        # Work description — merged block starts at F (col6) on the date row
        ws.cell(row=wr, column=6).value = d.get("description", "")

    # ── Additional Comments → F47  (merged F47:AA47) ──────────────────────────
    ws.cell(row=47, column=6).value = data.get("additional_comments", "")

    out = io.BytesIO()
    wb.save(out)
    out.seek(0)
    return out.read()


def send_email_report(smtp_host, smtp_port, smtp_user, smtp_pass, recipients, subject, body, attachment, filename):
    msg = MIMEMultipart()
    msg["From"] = smtp_user
    msg["To"] = ", ".join(recipients)
    msg["Subject"] = subject
    msg.attach(MIMEText(body, "plain"))
    part = MIMEBase("application", "octet-stream")
    part.set_payload(attachment)
    encoders.encode_base64(part)
    part.add_header("Content-Disposition", f'attachment; filename="{filename}"')
    msg.attach(part)
    with smtplib.SMTP(smtp_host, smtp_port) as s:
        s.starttls(); s.login(smtp_user, smtp_pass)
        s.sendmail(smtp_user, recipients, msg.as_string())


# ─── Session state defaults ────────────────────────────────────────────────────
_day_default = lambda: {"active":False,"straight":0.0,"overtime":0.0,"doubletime":0.0,
                         "travel_time":0.0,"working":0.0,"wait":0.0,"description":""}
for k, v in [("step",1),("days",[_day_default() for _ in range(7)]),
             ("generated_bytes",None),("generated_filename",""),("additional_comments","")]:
    if k not in st.session_state:
        st.session_state[k] = v

step = st.session_state.step

# ─── Compute current filename for topbar export button ────────────────────────
_fname_live = generate_filename(
    st.session_state.get("project_num",""),
    st.session_state.get("start_date", date.today()),
    st.session_state.get("emp_no",""),
    st.session_state.get("last_name",""),
    st.session_state.get("report_type","Other (O)")
)

# ─── TOP BAR ──────────────────────────────────────────────────────────────────
st.markdown(f"""
<div class="topbar">
  <div>{FANUC_LOGO_SVG}</div>
  <div class="topbar-right" onclick="void(0)">
    ⬇ &nbsp;EXPORT REPORT
  </div>
</div>
""", unsafe_allow_html=True)

ad_slot(ADSENSE_SLOT_TOP, "60px")

# ─── Step pills ───────────────────────────────────────────────────────────────
pills = [("1","Project"),("2","Engineer"),("3","Days"),("4","Generate"),("5","Email")]
html = '<div class="steps-wrap"><div class="steps">' + "".join(
    f'<div class="sp {"on" if int(n)==step else ("done" if int(n)<step else "")}">{n} · {lbl}</div>'
    for n, lbl in pills
) + '</div></div>'
st.markdown(html, unsafe_allow_html=True)


# ══════════════════════════════════════════════════════════════════
# STEP 1 – PROJECT INFO
# ══════════════════════════════════════════════════════════════════
if step == 1:
    st.markdown("""
    <div class="page-title">
      <span class="icon">🏭</span>
      <h2>Project &amp;<br>Customer Details</h2>
    </div>
    """, unsafe_allow_html=True)

    st.markdown('<div class="card">', unsafe_allow_html=True)
    project_num          = st.text_input("Project #",               value=st.session_state.get("project_num",""),          placeholder="e.g. 3011187")
    customer_contact     = st.text_input("Customer Name",           value=st.session_state.get("customer_contact",""),      placeholder="e.g. General Motors")
    plant_location       = st.text_input("Plant / Location",        value=st.session_state.get("plant_location",""),        placeholder="e.g. GM Lansing Delta Assembly")
    program_process_cell = st.text_input("Program / Process / Cell",value=st.session_state.get("program_process_cell",""), placeholder="e.g. Body Shop / Cell 12")
    phone                = st.text_input("Customer Phone #",        value=st.session_state.get("phone",""),                 placeholder="e.g. 555-123-4567")
    report_type          = st.selectbox("Report Type",              list(REPORT_TYPES.keys()),
                                         index=list(REPORT_TYPES.keys()).index(st.session_state.get("report_type","Other (O)")))
    col_exp, col_rep = st.columns([2, 1])
    with col_exp:
        expense_amount = st.number_input("Expense Amount ($)", value=float(st.session_state.get("expense_amount", 0)), min_value=0.0, step=0.01, format="%.2f")
    with col_rep:
        exp_rep = st.selectbox("Exp Rep?", ["No", "Yes"],
                                index=0 if not st.session_state.get("exp_rep", False) else 1)
    st.markdown('</div>', unsafe_allow_html=True)

    st.markdown('<div class="bottom-nav">', unsafe_allow_html=True)
    if st.button("Continue →", use_container_width=True):
        if not project_num:
            st.error("Project # is required.")
        else:
            st.session_state.update(dict(project_num=project_num, plant_location=plant_location,
                report_type=report_type, customer_contact=customer_contact, phone=phone,
                program_process_cell=program_process_cell, expense_amount=expense_amount,
                exp_rep=(exp_rep == "Yes"), step=2))
            st.rerun()
    st.markdown('</div>', unsafe_allow_html=True)


# ══════════════════════════════════════════════════════════════════
# STEP 2 – ENGINEER DETAILS
# ══════════════════════════════════════════════════════════════════
elif step == 2:
    st.markdown("""
    <div class="page-title">
      <span class="icon">👷</span>
      <h2>Engineer<br>Details</h2>
    </div>
    """, unsafe_allow_html=True)

    st.markdown('<div class="card">', unsafe_allow_html=True)
    engineer_first_name = st.text_input("First Name",       value=st.session_state.get("engineer_first_name",""), placeholder="e.g. Joy")
    last_name           = st.text_input("Last Name",         value=st.session_state.get("last_name",""),           placeholder="e.g. Smith")
    engineer_print_name = st.text_input("Full Print Name",  value=st.session_state.get("engineer_print_name",""), placeholder="e.g. Joy Smith")
    emp_no              = st.text_input("Employee # (optional — uses Last Name if blank)",
                                         value=st.session_state.get("emp_no",""), placeholder="e.g. 260216")
    start_date          = st.date_input("Start Date (Week of)", value=st.session_state.get("start_date", date.today()))
    st.markdown('</div>', unsafe_allow_html=True)

    # Live filename preview
    fname = generate_filename(
        st.session_state.get("project_num",""), start_date,
        emp_no, last_name, st.session_state.get("report_type","Other (O)"))
    st.markdown(f"""
    <div style="padding:0 12px;">
      <div class="fname-label">Auto-generated filename</div>
      <div class="fname">{fname}</div>
    </div>
    """, unsafe_allow_html=True)

    st.markdown('<div class="bottom-nav" style="margin-top:12px;">', unsafe_allow_html=True)
    c1, c2 = st.columns(2)
    with c1:
        if st.button("← Back", use_container_width=True):
            st.session_state.step = 1; st.rerun()
    with c2:
        if st.button("Continue →", use_container_width=True):
            if not engineer_first_name:
                st.error("First Name is required.")
            elif not emp_no and not last_name:
                st.error("Enter Employee # or Last Name.")
            else:
                st.session_state.update(dict(engineer_first_name=engineer_first_name,
                    last_name=last_name, engineer_print_name=engineer_print_name,
                    emp_no=emp_no, start_date=start_date, step=3))
                st.rerun()
    st.markdown('</div>', unsafe_allow_html=True)


# ══════════════════════════════════════════════════════════════════
# STEP 3 – DAILY WORK LOG
# ══════════════════════════════════════════════════════════════════
elif step == 3:
    sd   = st.session_state.get("start_date", date.today())
    days = st.session_state.days

    st.markdown("""
    <div class="page-title">
      <span class="icon">📅</span>
      <h2>Daily Work Log</h2>
      <p>Tap a day to expand and enter hours</p>
    </div>
    """, unsafe_allow_html=True)

    # Hour totals
    ts = sum(d.get("straight",0) or 0 for d in days if d.get("active"))
    to = sum(d.get("overtime",0) or 0 for d in days if d.get("active"))
    tt = sum(d.get("travel_time",0) or 0 for d in days if d.get("active"))
    total = ts + to + sum(d.get("doubletime",0) or 0 for d in days if d.get("active")) + tt

    c1,c2,c3,c4 = st.columns(4)
    c1.metric("S",  f"{ts:.1f}h")
    c2.metric("OT", f"{to:.1f}h")
    c3.metric("TT", f"{tt:.1f}h")
    c4.metric("Total", f"{total:.1f}h")

    ad_slot(ADSENSE_SLOT_MID, "80px")

    st.markdown('<div style="padding:0 12px;">', unsafe_allow_html=True)
    for i in range(7):
        d     = days[i]
        label = (sd + timedelta(days=i)).strftime("%A, %b %d")
        badge = "✅" if d.get("active") else "○"
        with st.expander(f"{badge}  Day {i+1} — {label}", expanded=False):
            active = st.checkbox("Include this day in report", value=d.get("active",False), key=f"act_{i}")
            st.session_state.days[i]["active"] = active
            if active:
                a, b, c_ = st.columns(3)
                with a:
                    v = st.number_input("Straight", 0.0, 24.0, float(d.get("straight") or 0), .5, key=f"s_{i}", format="%.1f")
                    st.session_state.days[i]["straight"] = v
                with b:
                    v = st.number_input("Overtime", 0.0, 24.0, float(d.get("overtime") or 0), .5, key=f"ot_{i}", format="%.1f")
                    st.session_state.days[i]["overtime"] = v
                with c_:
                    v = st.number_input("Double", 0.0, 24.0, float(d.get("doubletime") or 0), .5, key=f"dt_{i}", format="%.1f")
                    st.session_state.days[i]["doubletime"] = v

                a2, b2, c2_ = st.columns(3)
                with a2:
                    v = st.number_input("Travel", 0.0, 24.0, float(d.get("travel_time") or 0), .5, key=f"tt_{i}", format="%.1f")
                    st.session_state.days[i]["travel_time"] = v
                with b2:
                    v = st.number_input("Working", 0.0, 24.0, float(d.get("working") or 0), .5, key=f"w_{i}", format="%.1f")
                    st.session_state.days[i]["working"] = v
                with c2_:
                    v = st.number_input("Wait", 0.0, 24.0, float(d.get("wait") or 0), .5, key=f"wt_{i}", format="%.1f")
                    st.session_state.days[i]["wait"] = v

                desc = st.text_area("Work Description", value=d.get("description",""),
                                     placeholder="Describe work performed this day…",
                                     key=f"desc_{i}", height=90)
                st.session_state.days[i]["description"] = desc
    st.markdown('</div>', unsafe_allow_html=True)

    st.markdown('<div class="card" style="margin-top:8px;">', unsafe_allow_html=True)
    ac = st.text_area("Additional Comments", value=st.session_state.get("additional_comments",""),
                       placeholder="Any extra notes or follow-up items…", height=80)
    st.session_state["additional_comments"] = ac
    st.markdown('</div>', unsafe_allow_html=True)

    st.markdown('<div class="bottom-nav">', unsafe_allow_html=True)
    c1, c2 = st.columns(2)
    with c1:
        if st.button("← Back",       use_container_width=True): st.session_state.step=2; st.rerun()
    with c2:
        if st.button("Continue →",   use_container_width=True): st.session_state.step=4; st.rerun()
    st.markdown('</div>', unsafe_allow_html=True)


# ══════════════════════════════════════════════════════════════════
# STEP 4 – GENERATE & DOWNLOAD
# ══════════════════════════════════════════════════════════════════
elif step == 4:
    fname = generate_filename(
        st.session_state.get("project_num",""),
        st.session_state.get("start_date", date.today()),
        st.session_state.get("emp_no",""),
        st.session_state.get("last_name",""),
        st.session_state.get("report_type","Other (O)"))

    active_days = sum(1 for d in st.session_state.days if d.get("active"))
    total_hrs = sum(
        (d.get("straight") or 0)+(d.get("overtime") or 0)+
        (d.get("doubletime") or 0)+(d.get("travel_time") or 0)
        for d in st.session_state.days if d.get("active"))

    st.markdown("""
    <div class="page-title">
      <span class="icon">⚙️</span>
      <h2>Review &amp;<br>Export Report</h2>
    </div>
    """, unsafe_allow_html=True)

    st.markdown(f"""
    <div class="card">
      <div class="card-label">Report Summary</div>
      <div style="font-size:14px;color:#374151;line-height:2.2;">
        <b>Project:</b> {st.session_state.get('project_num','—')}<br>
        <b>Plant:</b> {st.session_state.get('plant_location','—')}<br>
        <b>Engineer:</b> {st.session_state.get('engineer_print_name','—')}
          &nbsp;·&nbsp; #{st.session_state.get('emp_no', st.session_state.get('last_name','—'))}<br>
        <b>Active Days:</b> {active_days} &nbsp;·&nbsp; <b>Total Hours:</b> {total_hrs:.1f}
      </div>
      <div class="fname-label" style="margin-top:14px;">Generated Filename</div>
      <div class="fname">{fname}</div>
    </div>
    """, unsafe_allow_html=True)

    st.markdown('<div class="bottom-nav">', unsafe_allow_html=True)
    if st.button("⚙️  GENERATE FIR REPORT", use_container_width=True):
        with st.spinner("Building your FIR Excel file…"):
            try:
                xls = populate_fir_excel({
                    "project_num":        st.session_state.get("project_num",""),
                    "plant_location":     st.session_state.get("plant_location",""),
                    "customer_contact":   st.session_state.get("customer_contact",""),
                    "phone":              st.session_state.get("phone",""),
                    "report_type":        st.session_state.get("report_type","Other (O)"),
                    "expense_amount":     st.session_state.get("expense_amount",0),
                    "exp_rep":            st.session_state.get("exp_rep", False),
                    "program_process_cell": st.session_state.get("program_process_cell",""),
                    "engineer_first_name":st.session_state.get("engineer_first_name",""),
                    "engineer_print_name":st.session_state.get("engineer_print_name",""),
                    "emp_no":             st.session_state.get("emp_no",""),
                    "last_name":          st.session_state.get("last_name",""),
                    "start_date":         st.session_state.get("start_date", date.today()),
                    "days":               st.session_state.days,
                    "additional_comments":st.session_state.get("additional_comments",""),
                })
                st.session_state.generated_bytes    = xls
                st.session_state.generated_filename = fname
                st.success("✅ Report ready!")
            except Exception as e:
                st.error(f"Error generating report: {e}")

    if st.session_state.generated_bytes:
        st.download_button(
            "⬇️  DOWNLOAD  " + st.session_state.generated_filename,
            data=st.session_state.generated_bytes,
            file_name=st.session_state.generated_filename,
            mime="application/vnd.ms-excel.sheet.macroEnabled.12",
            use_container_width=True)
        ad_slot(ADSENSE_SLOT_BOTTOM, "90px")

    c1, c2 = st.columns(2)
    with c1:
        if st.button("← Back",         use_container_width=True): st.session_state.step=3; st.rerun()
    with c2:
        if st.button("📧 Email Report →", use_container_width=True): st.session_state.step=5; st.rerun()
    st.markdown('</div>', unsafe_allow_html=True)


# ══════════════════════════════════════════════════════════════════
# STEP 5 – EMAIL
# ══════════════════════════════════════════════════════════════════
elif step == 5:
    st.markdown("""
    <div class="page-title">
      <span class="icon">📧</span>
      <h2>Send Report<br>by Email</h2>
    </div>
    """, unsafe_allow_html=True)

    if not st.session_state.generated_bytes:
        st.warning("⚠️ Please generate the report first (Step 4).")
    else:
        st.markdown(f"""
        <div class="card">
          <div class="card-label">Attachment</div>
          <div style="font-size:13px;color:#374151;word-break:break-all;">
            📎 {st.session_state.generated_filename}
          </div>
        </div>
        """, unsafe_allow_html=True)

        st.markdown('<div class="card">', unsafe_allow_html=True)
        smtp_host      = st.text_input("SMTP Host",         value="smtp.gmail.com")
        smtp_port      = st.number_input("Port",            value=587, min_value=1, max_value=65535)
        smtp_user      = st.text_input("Your Email",        placeholder="you@gmail.com")
        smtp_pass      = st.text_input("App Password",      type="password",
                                        help="Gmail: Account → Security → App Passwords")
        recipients_raw = st.text_area("Recipients (one per line)",
                                       placeholder="manager@fanuc.com\ncustomer@plant.com", height=90)
        subject        = st.text_input("Subject",
                                        value=f"FIR – Project {st.session_state.get('project_num','')} – {st.session_state.get('engineer_print_name','')}")
        st.markdown('</div>', unsafe_allow_html=True)

        st.markdown('<div class="bottom-nav">', unsafe_allow_html=True)
        if st.button("📤  SEND EMAIL", use_container_width=True):
            if not all([smtp_host, smtp_user, smtp_pass, recipients_raw]):
                st.error("Please fill in all email fields.")
            else:
                recips = [r.strip() for r in recipients_raw.strip().splitlines() if r.strip()]
                body = (f"FIR attached — Project #{st.session_state.get('project_num','')}\n\n"
                        f"Engineer: {st.session_state.get('engineer_print_name','')}\n"
                        f"Employee ID: {st.session_state.get('emp_no', st.session_state.get('last_name',''))}\n"
                        f"Week of: {st.session_state.get('start_date', date.today()).strftime('%B %d, %Y')}\n\n"
                        "Sent via FANUC FIR Automation System.")
                with st.spinner("Sending…"):
                    try:
                        send_email_report(smtp_host, int(smtp_port), smtp_user, smtp_pass,
                                          recips, subject, body,
                                          st.session_state.generated_bytes,
                                          st.session_state.generated_filename)
                        st.success(f"✅ Sent to: {', '.join(recips)}")
                    except Exception as e:
                        st.error(f"Failed to send: {e}")
        st.markdown('</div>', unsafe_allow_html=True)

    st.markdown('<div class="bottom-nav" style="margin-top:10px;">', unsafe_allow_html=True)
    c1, c2 = st.columns(2)
    with c1:
        if st.button("← Back",        use_container_width=True): st.session_state.step=4; st.rerun()
    with c2:
        if st.button("🔄 New Report", use_container_width=True):
            for k in ["project_num","plant_location","report_type","customer_contact","phone",
                      "program_process_cell","expense_amount","engineer_first_name","last_name",
                      "engineer_print_name","emp_no","start_date","additional_comments",
                      "generated_bytes","generated_filename"]:
                st.session_state.pop(k, None)
            st.session_state.days = [_day_default() for _ in range(7)]
            st.session_state.step = 1
            st.rerun()
    st.markdown('</div>', unsafe_allow_html=True)

# ─── Footer ───────────────────────────────────────────────────────────────────
st.markdown("""
<div style="text-align:center;padding:28px 0 10px;color:#9CA3AF;font-size:10px;letter-spacing:1px;">
  FANUC AMERICA CORPORATION · FIR SYSTEM v2.0<br>
  <span style="color:#D1D5DB;">Not for external distribution</span>
</div>
""", unsafe_allow_html=True)
